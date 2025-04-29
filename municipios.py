from openpyxl import load_workbook
import os
import glob, os

# Colocar o diretório onde estão os arquivos que você quer copiar
os.chdir("C:/Users/André/Desktop/Coisas do André/UFMG/Pesquisa/OSS/Pesquisa municipios/2_Negativos/BA")
ox = []
for file in glob.glob("*.*"):
    for j in range(0, 1):  # Pra copiar o nome do arquivo apenas 1 vez
        replaced_text = file.replace('.xlsx', '') # Retira o final do arquivo se for ".xlsx".
                                                            # Se for pdf tem que ser feito manualmente ou mudar essa linha
        ox.append(replaced_text)

# Colocar aqui o atalho da planilha excel (deve ser criada previamente)
file_dir = 'C:/Users/André/Desktop/Coisas do André/UFMG/Pesquisa/OSS/Pesquisa municipios/municipios.xlsx'
file1 = load_workbook(filename = file_dir)

# Colocar aqui o nome da aba no excel para onde se deseja copiar o nome dos arquivos
sheet1 = file1['NegativosBA']
last_row = 1

for counter, item in enumerate(ox):
    sheet1.cell(row=(last_row + counter), column=1).value = item

# Salvar o arquivo
file1.save(file_dir)

# Mudar o nome do diretorio e aba a cada vez que quiser rodar o codigo novamente
# Deve ter um jeito mais fácil de fazer isso, mas meu conhecimento em python nao permite :/